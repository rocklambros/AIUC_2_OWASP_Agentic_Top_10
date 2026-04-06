"""Output generation: Excel (5-sheet workbook) + JSON mapping file (v2 schema).

The workbook contains a README instructional sheet followed by 4 data sheets.
The JSON output conforms to schemas/crosswalk-mapping-v2.schema.json.
"""

from __future__ import annotations

import json
import logging
import re
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from aiuc.models import (
    ClassificationAccuracy,
    ClassificationConfig,
    ConfidenceTier,
    CrosswalkMetadata,
    CrosswalkOutput,
    CrosswalkSummary,
    GapAnalysis,
    MappingOutput,
    OwaspToSourceMapping,
    PipelineConfig,
    PipelineSignal,
    PipelineThresholds,
    RationaleCode,
    RelevanceLevel,
    SourceControlEntry,
    SourceToOwaspMapping,
    StandardInfo,
    UnmappedControl,
    V2ControlLevel,
    V2OwaspEntry,
)
from aiuc.taxonomy import AIUC_CONTROL_FUNCTIONS, RATIONALE_LABELS

logger = logging.getLogger(__name__)

# ── Style constants ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
DIRECT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RELATED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PRIMARY_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PRIMARY_FONT = Font(color="FFFFFF")
SECONDARY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# README-specific styles
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=13, color="1F4E79")
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
BODY_FONT = Font(size=11)
BODY_WRAP = Alignment(wrap_text=True, vertical="top")
TABLE_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TABLE_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def _style_header(ws: Worksheet, num_cols: int) -> None:
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws: Worksheet, num_cols: int, max_rows: int) -> None:
    """Auto-fit column widths."""
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in range(1, min(max_rows + 1, 100)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        adjusted = min(max_len + 2, 50)
        ws.column_dimensions[get_column_letter(col)].width = adjusted


def _apply_confidence_fill(ws: Worksheet, col: int, start_row: int, end_row: int) -> None:
    """Color-code confidence tier cells."""
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col)
        val = str(cell.value) if cell.value else ""
        if val == ConfidenceTier.DIRECT.value:
            cell.fill = DIRECT_FILL
        elif val == ConfidenceTier.RELATED.value:
            cell.fill = RELATED_FILL
        cell.border = THIN_BORDER


def _apply_relevance_fill(ws: Worksheet, col: int, start_row: int, end_row: int) -> None:
    """Color-code relevance cells: Primary=blue, Secondary=gray."""
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col)
        val = str(cell.value) if cell.value else ""
        if val == "Primary":
            cell.fill = PRIMARY_FILL
            cell.font = PRIMARY_FONT
        elif val == "Secondary":
            cell.fill = SECONDARY_FILL
        cell.border = THIN_BORDER


# ── README Sheet ─────────────────────────────────────────────────────────────


def _readme_section_header(ws: Worksheet, row: int, title: str) -> int:
    """Write a section header spanning columns A-D and return the next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER
    for col in range(2, 5):
        ws.cell(row=row, column=col).border = THIN_BORDER
    return row + 1


def _readme_body_row(ws: Worksheet, row: int, text: str) -> int:
    """Write a body text row merged across A-D and return the next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = BODY_FONT
    cell.alignment = BODY_WRAP
    return row + 1


def _readme_table_header(ws: Worksheet, row: int, headers: list[str]) -> int:
    """Write a table header row and return the next row."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = TABLE_HEADER_FILL
        cell.font = TABLE_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    return row + 1


def _readme_table_row(ws: Worksheet, row: int, values: list[str]) -> int:
    """Write a table data row and return the next row."""
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = BODY_FONT
        cell.alignment = BODY_WRAP
        cell.border = THIN_BORDER
    return row + 1


def _write_readme_sheet(wb: Workbook) -> None:
    """Write the README instructional sheet as the first tab."""
    ws = wb.create_sheet("README", 0)
    num_cols = 4
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40

    row = 1

    # ── Title ────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value="AIUC-1 \u2194 OWASP Agentic Top 10 Mapping Workbook")
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    # ── Workbook Overview ────────────────────────────────────────────────
    row = _readme_section_header(ws, row, "Workbook Overview")
    row = _readme_body_row(ws, row, (
        "This workbook provides a bi-directional mapping between AIUC-1 (AI Use Case "
        "Controls) and the OWASP Top 10 for Agentic Applications (2026). Each mapping "
        "is annotated with a rationale code (why the control is relevant), a relevance "
        "classification (Primary or Secondary), and a function class (the control's "
        "defense category)."
    ))
    row = _readme_body_row(ws, row, (
        "The workbook contains 5 sheets: this README plus 4 data sheets covering "
        "control-level and activity-level mappings in both directions."
    ))
    row += 1

    # ── Rationale Taxonomy ───────────────────────────────────────────────
    row = _readme_section_header(ws, row, "Rationale Taxonomy (Function Classes)")
    row = _readme_body_row(ws, row, (
        "Each control is assigned a function class describing its defense role. "
        "The same code appears as 'Function Class' (a property of the control) "
        "and 'Rationale' (why the mapping exists)."
    ))
    row = _readme_table_header(ws, row, ["Code", "Label", "Definition", ""])
    taxonomy_rows = [
        ["PREV", "Prevent", "Directly blocks the core attack mechanism", ""],
        ["SCOPE", "Constrain scope", "Limits blast radius after compromise", ""],
        ["GATE", "Human gate", "Enforces human approval or intervention", ""],
        ["DETECT", "Detect and trace", "Provides runtime detection or forensic traceability", ""],
        ["VALID", "Validate and test",
         "Tests or audits that other controls work", ""],
        ["GOVERN", "Policy and governance",
         "Establishes organizational policy or accountability", ""],
        ["ISOLATE", "Isolate and contain",
         "Enforces architectural separation", ""],
        ["DISCLOSE", "Disclose and calibrate",
         "Provides transparency to calibrate trust", ""],
    ]
    for t in taxonomy_rows:
        row = _readme_table_row(ws, row, t)
    row += 1

    # ── Relevance Levels ─────────────────────────────────────────────────
    row = _readme_section_header(ws, row, "Relevance Classification")
    row = _readme_body_row(ws, row, (
        "Each mapping is classified as Primary or Secondary based on the "
        "threat context, not the rationale code alone."
    ))
    row = _readme_table_header(ws, row, ["Level", "Color", "Meaning", ""])
    prim_row = row
    row = _readme_table_row(ws, row, [
        "Primary", "Blue", "Directly mitigates the core risk", "",
    ])
    ws.cell(row=prim_row, column=2).fill = PRIMARY_FILL
    ws.cell(row=prim_row, column=2).font = PRIMARY_FONT
    sec_row = row
    row = _readme_table_row(ws, row, [
        "Secondary", "Gray", "Addresses a related consequence or provides supporting control", "",
    ])
    ws.cell(row=sec_row, column=2).fill = SECONDARY_FILL
    row += 1

    # ── Tab-by-Tab Guide ─────────────────────────────────────────────────
    row = _readme_section_header(ws, row, "Tab-by-Tab Guide")

    row = _readme_body_row(ws, row, "\u25B6 Sheet: AIUC\u2192OWASP (Control)")
    row = _readme_body_row(ws, row, (
        "Shows which OWASP agentic risks each AIUC control addresses, with "
        "function class, rationale, relevance, score, and sub-scores."
    ))
    row = _readme_table_header(ws, row, ["Column", "Definition", "", ""])
    for c in [
        ["AIUC ID", "AIUC-1 control identifier"],
        ["AIUC Title", "Name of the AIUC control"],
        ["AIUC Domain", "Domain category (Data & Privacy, Security, Safety, etc.)"],
        ["Function Class", "Control's defense function (PREV, SCOPE, GATE, etc.)"],
        ["OWASP ID", "Matched OWASP entry identifier (ASI01\u2013ASI10)"],
        ["OWASP Title", "Name of the matched OWASP risk"],
        ["Rationale", "Why this mapping exists (same as Function Class)"],
        ["Rationale Label", "Human-readable label for the rationale code"],
        ["Relevance", "Primary (direct mitigation) or Secondary (supporting)"],
        ["Score", "Composite similarity score (0.0\u20131.0)"],
        ["Confidence", "Direct (\u22650.55) or Related (\u22650.28)"],
        ["Ref Bridge", "Reference Bridge sub-score"],
        ["Semantic", "Semantic similarity sub-score"],
        ["Keyword", "Keyword overlap sub-score"],
        ["Relationship", "How the control relates (Prevents, Mitigates, etc.)"],
    ]:
        row = _readme_table_row(ws, row, [c[0], c[1], "", ""])
    row += 1

    row = _readme_body_row(ws, row, "\u25B6 Sheet: OWASP\u2192AIUC (Control)")
    row = _readme_body_row(ws, row, (
        "Shows which AIUC controls address each OWASP risk, with function "
        "coverage analysis and per-mapping rationale/relevance."
    ))
    row = _readme_table_header(ws, row, ["Column", "Definition", "", ""])
    for c in [
        ["OWASP ID", "OWASP entry identifier"],
        ["OWASP Title", "Name of the OWASP risk"],
        ["Rank", "OWASP ranking (1 = highest priority)"],
        ["Function Coverage", "Rationale codes covered (e.g., SCOPE:3, PREV:2)"],
        ["Uncovered Functions", "Rationale codes with zero mapped controls"],
        ["AIUC ID", "Matched AIUC control identifier"],
        ["AIUC Title", "Name of the matched AIUC control"],
        ["AIUC Domain", "Domain category of the matched control"],
        ["Function Class", "Control's defense function"],
        ["Rationale", "Rationale code for this mapping"],
        ["Rationale Label", "Human-readable rationale label"],
        ["Relevance", "Primary or Secondary"],
        ["Score", "Composite similarity score"],
        ["Confidence", "Direct or Related"],
        ["Ref Bridge", "Reference Bridge sub-score"],
        ["Semantic", "Semantic similarity sub-score"],
        ["Keyword", "Keyword overlap sub-score"],
        ["Relationship", "How the control relates to the risk"],
    ]:
        row = _readme_table_row(ws, row, [c[0], c[1], "", ""])
    row += 1

    row = _readme_body_row(ws, row, "\u25B6 Sheet: AIUC\u2192OWASP (Activity)")
    row = _readme_body_row(ws, row, (
        "Granular activity-level view mapping individual AIUC sub-activities to OWASP risks."
    ))
    row += 1

    row = _readme_body_row(ws, row, "\u25B6 Sheet: OWASP\u2192AIUC (Activity)")
    row = _readme_body_row(ws, row, (
        "Reverse activity-level view showing which AIUC activities address each OWASP risk."
    ))
    row += 1

    # ── Confidence Tiers ─────────────────────────────────────────────────
    row = _readme_section_header(ws, row, "Understanding Confidence Tiers")
    row = _readme_table_header(ws, row, ["Tier", "Color", "Score Range", "Meaning"])
    direct_row = row
    row = _readme_table_row(ws, row, [
        "Direct", "Green", "\u2265 0.55",
        "Strong, well-supported connection.",
    ])
    ws.cell(row=direct_row, column=2).fill = DIRECT_FILL
    related_row = row
    row = _readme_table_row(ws, row, [
        "Related", "Yellow", "\u2265 0.28",
        "Meaningful but less direct connection.",
    ])
    ws.cell(row=related_row, column=2).fill = RELATED_FILL
    row = _readme_table_row(ws, row, [
        "Governance Floor", "", "\u2265 0.22",
        "GOVERN/DISCLOSE controls with function match promoted to Related.",
    ])
    row += 1

    # ── Scoring Methodology ──────────────────────────────────────────────
    row = _readme_section_header(ws, row, "Scoring Methodology")
    row = _readme_body_row(ws, row, (
        "Composite Score = 0.35 \u00D7 Reference Bridge + 0.25 \u00D7 Semantic "
        "+ 0.15 \u00D7 Keyword + 0.25 \u00D7 Function Match"
    ))
    row = _readme_table_header(ws, row, ["Signal", "Weight", "Method", "What It Measures"])
    for s in [
        ["Reference Bridge", "35%",
         "Jaccard overlap on shared OWASP LLM Top 10 references",
         "Shared vulnerability references between both frameworks."],
        ["Semantic Similarity", "25%",
         "Sentence-transformer embeddings (all-MiniLM-L6-v2)",
         "Meaning similarity via NLP, with prevention-guideline boosting."],
        ["TF-IDF Keyword", "15%",
         "TF-IDF cosine with domain-specific synonym expansion",
         "Explicit term matches with AI-security synonym groups."],
        ["Function Match", "25%",
         "Binary match: control function class \u2208 threat function profile",
         "Whether the control's defense category is relevant to the threat."],
    ]:
        row = _readme_table_row(ws, row, s)
    row += 1

    # ── Framework References ─────────────────────────────────────────────
    row = _readme_section_header(ws, row, "AIUC-1 Domains Quick Reference")
    row = _readme_table_header(ws, row, ["Domain", "ID Range", "Focus", ""])
    for d in [
        ["A \u2014 Data & Privacy", "A001\u2013A007", "Data leakage, PII, IP rights"],
        ["B \u2014 Security", "B001\u2013B009", "Adversarial robustness, access control"],
        ["C \u2014 Safety", "C001\u2013C012", "Harmful outputs, vulnerability prevention"],
        ["D \u2014 Reliability", "D001\u2013D004", "Hallucinations, unsafe tool calls"],
        ["E \u2014 Accountability", "E001\u2013E017", "Vendor management, audit, compliance"],
        ["F \u2014 Society", "F001\u2013F002", "Societal impact, responsible disclosure"],
    ]:
        row = _readme_table_row(ws, row, [d[0], d[1], d[2], ""])
    row += 1

    row = _readme_section_header(ws, row, "OWASP Agentic Top 10 (2026) Quick Reference")
    row = _readme_table_header(ws, row, ["Rank", "ID", "Title", "Description"])
    for e in [
        ["1", "ASI01", "Agent Goal Hijack",
         "Attackers manipulate an agent's objectives through prompt injection."],
        ["2", "ASI02", "Tool Misuse and Exploitation",
         "Agents misuse legitimate tools due to injection or misalignment."],
        ["3", "ASI03", "Identity & Privilege Abuse",
         "Agents inherit or escalate privileges beyond what is needed."],
        ["4", "ASI04", "Agentic Supply Chain Vulnerabilities",
         "Compromised tools or dependencies introduce malicious capabilities."],
        ["5", "ASI05", "Unexpected Code Execution",
         "Agents generate and execute unintended or malicious code."],
        ["6", "ASI06", "Memory & Context Poisoning",
         "Persistent corruption of an agent's stored context."],
        ["7", "ASI07", "Multi-Agent Trust Issues",
         "Agents trust peer messages without verification."],
        ["8", "ASI08", "Insufficient Monitoring & Logging",
         "Lack of comprehensive logging hinders detection."],
        ["9", "ASI09", "Lack of Human Oversight",
         "Agents operate without adequate human review gates."],
        ["10", "ASI10", "Rogue Agents",
         "Agents develop autonomous misalignment from intended objectives."],
    ]:
        row = _readme_table_row(ws, row, e)

    ws.print_area = f"A1:D{row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ── Sheet 1: AIUC -> OWASP (Control) ──────────────────────────────────────


def _write_aiuc_to_owasp_control(wb: Workbook, mapping: MappingOutput) -> None:
    """Write Sheet 1: AIUC -> OWASP at control level."""
    ws = wb.create_sheet("AIUC\u2192OWASP (Control)")

    headers = [
        "AIUC ID", "AIUC Title", "AIUC Domain", "Function Class",
        "OWASP ID", "OWASP Title", "Rationale", "Rationale Label", "Relevance",
        "Score", "Confidence",
        "Ref Bridge", "Semantic", "Keyword", "Relationship",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, len(headers))

    row = 2
    for ctrl in mapping.control_level.aiuc_to_owasp:
        func_class = AIUC_CONTROL_FUNCTIONS.get(ctrl.aiuc_id, "GOVERN")
        if not ctrl.mappings:
            ws.cell(row=row, column=1, value=ctrl.aiuc_id)
            ws.cell(row=row, column=2, value=ctrl.aiuc_title)
            ws.cell(row=row, column=3, value=ctrl.aiuc_domain)
            ws.cell(row=row, column=4, value=func_class)
            ws.cell(row=row, column=5, value="\u2014")
            ws.cell(row=row, column=11, value="None")
            row += 1
            continue

        for m in ctrl.mappings:
            ws.cell(row=row, column=1, value=ctrl.aiuc_id)
            ws.cell(row=row, column=2, value=ctrl.aiuc_title)
            ws.cell(row=row, column=3, value=ctrl.aiuc_domain)
            ws.cell(row=row, column=4, value=func_class)
            ws.cell(row=row, column=5, value=m.owasp_id)
            ws.cell(row=row, column=6, value=m.owasp_title)
            ws.cell(row=row, column=7, value=m.rationale_code.value)
            ws.cell(row=row, column=8, value=m.rationale_label)
            ws.cell(row=row, column=9, value=m.relevance.value)
            ws.cell(row=row, column=10, value=m.score)
            ws.cell(row=row, column=11, value=m.confidence.value)
            ws.cell(row=row, column=12, value=m.signals.reference_bridge if m.signals else 0.0)
            ws.cell(row=row, column=13, value=m.signals.semantic if m.signals else 0.0)
            ws.cell(row=row, column=14, value=m.signals.keyword if m.signals else 0.0)
            ws.cell(row=row, column=15, value=m.relationship_type.value)
            row += 1

    _apply_confidence_fill(ws, 11, 2, row - 1)
    _apply_relevance_fill(ws, 9, 2, row - 1)
    _auto_width(ws, len(headers), row - 1)


# ── Sheet 2: OWASP -> AIUC (Control) ──────────────────────────────────────


def _compute_coverage(
    entry_mappings: list,
) -> tuple[str, str]:
    """Compute function coverage string and uncovered functions string."""
    coverage: dict[str, int] = {}
    for m in entry_mappings:
        rc = m.rationale_code
        code = rc.value if hasattr(rc, "value") else str(rc)
        coverage[code] = coverage.get(code, 0) + 1

    all_codes = {rc.value for rc in RationaleCode}
    uncovered = sorted(all_codes - set(coverage.keys()))

    # Format: "SCOPE:3, PREV:2, DETECT:1" sorted by count desc
    coverage_str = ", ".join(
        f"{k}:{v}" for k, v in sorted(coverage.items(), key=lambda x: -x[1])
    )
    uncovered_str = ", ".join(uncovered) if uncovered else "\u2014"
    return coverage_str or "\u2014", uncovered_str


def _write_owasp_to_aiuc_control(wb: Workbook, mapping: MappingOutput) -> None:
    """Write Sheet 2: OWASP -> AIUC at control level."""
    ws = wb.create_sheet("OWASP\u2192AIUC (Control)")

    headers = [
        "OWASP ID", "OWASP Title", "Rank",
        "Function Coverage", "Uncovered Functions",
        "AIUC ID", "AIUC Title", "AIUC Domain", "Function Class",
        "Rationale", "Rationale Label", "Relevance",
        "Score", "Confidence",
        "Ref Bridge", "Semantic", "Keyword", "Relationship",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, len(headers))

    row = 2
    for entry in mapping.control_level.owasp_to_aiuc:
        cov_str, uncov_str = _compute_coverage(entry.mappings)

        if not entry.mappings:
            ws.cell(row=row, column=1, value=entry.owasp_id)
            ws.cell(row=row, column=2, value=entry.owasp_title)
            ws.cell(row=row, column=3, value=entry.owasp_rank)
            ws.cell(row=row, column=4, value="\u2014")
            ws.cell(row=row, column=5, value=uncov_str)
            ws.cell(row=row, column=6, value="\u2014")
            ws.cell(row=row, column=14, value="None")
            row += 1
            continue

        for m in entry.mappings:
            ws.cell(row=row, column=1, value=entry.owasp_id)
            ws.cell(row=row, column=2, value=entry.owasp_title)
            ws.cell(row=row, column=3, value=entry.owasp_rank)
            ws.cell(row=row, column=4, value=cov_str)
            ws.cell(row=row, column=5, value=uncov_str)
            ws.cell(row=row, column=6, value=m.aiuc_id)
            ws.cell(row=row, column=7, value=m.aiuc_title)
            ws.cell(row=row, column=8, value=m.aiuc_domain)
            func_class = AIUC_CONTROL_FUNCTIONS.get(m.aiuc_id, "GOVERN")
            ws.cell(row=row, column=9, value=func_class)
            ws.cell(row=row, column=10, value=m.rationale_code.value)
            ws.cell(row=row, column=11, value=m.rationale_label)
            ws.cell(row=row, column=12, value=m.relevance.value)
            ws.cell(row=row, column=13, value=m.score)
            ws.cell(row=row, column=14, value=m.confidence.value)
            ws.cell(row=row, column=15, value=m.signals.reference_bridge if m.signals else 0.0)
            ws.cell(row=row, column=16, value=m.signals.semantic if m.signals else 0.0)
            ws.cell(row=row, column=17, value=m.signals.keyword if m.signals else 0.0)
            ws.cell(row=row, column=18, value=m.relationship_type.value)
            row += 1

    _apply_confidence_fill(ws, 14, 2, row - 1)
    _apply_relevance_fill(ws, 12, 2, row - 1)
    _auto_width(ws, len(headers), row - 1)


# ── Sheet 3: AIUC -> OWASP (Sub-Activity) ─────────────────────────────────


def _write_aiuc_to_owasp_sub(wb: Workbook, mapping: MappingOutput) -> None:
    """Write Sheet 3: AIUC -> OWASP at sub-activity level."""
    ws = wb.create_sheet("AIUC\u2192OWASP (Activity)")

    headers = [
        "Activity ID", "Parent Control", "Description",
        "OWASP ID", "Score", "Confidence",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, len(headers))

    row = 2
    for act in mapping.sub_activity_level.aiuc_to_owasp:
        if not act.mappings:
            ws.cell(row=row, column=1, value=act.activity_id)
            ws.cell(row=row, column=2, value=act.parent_control)
            ws.cell(row=row, column=3, value=act.description[:100])
            ws.cell(row=row, column=4, value="\u2014")
            ws.cell(row=row, column=6, value="None")
            row += 1
            continue

        for m in act.mappings:
            ws.cell(row=row, column=1, value=act.activity_id)
            ws.cell(row=row, column=2, value=act.parent_control)
            ws.cell(row=row, column=3, value=act.description[:100])
            ws.cell(row=row, column=4, value=m.owasp_id)
            ws.cell(row=row, column=5, value=m.score)
            ws.cell(row=row, column=6, value=m.confidence.value)
            row += 1

    _apply_confidence_fill(ws, 6, 2, row - 1)
    _auto_width(ws, len(headers), row - 1)


# ── Sheet 4: OWASP -> AIUC (Sub-Activity) ─────────────────────────────────


def _write_owasp_to_aiuc_sub(wb: Workbook, mapping: MappingOutput) -> None:
    """Write Sheet 4: OWASP -> AIUC at sub-activity level."""
    ws = wb.create_sheet("OWASP\u2192AIUC (Activity)")

    headers = [
        "OWASP ID", "OWASP Title",
        "Activity ID", "Parent Control", "Score", "Confidence",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, len(headers))

    row = 2
    for entry in mapping.sub_activity_level.owasp_to_aiuc:
        if not entry.mappings:
            ws.cell(row=row, column=1, value=entry.owasp_id)
            ws.cell(row=row, column=2, value=entry.owasp_title)
            ws.cell(row=row, column=3, value="\u2014")
            ws.cell(row=row, column=6, value="None")
            row += 1
            continue

        for m in entry.mappings:
            ws.cell(row=row, column=1, value=entry.owasp_id)
            ws.cell(row=row, column=2, value=entry.owasp_title)
            ws.cell(row=row, column=3, value=m.activity_id)
            ws.cell(row=row, column=4, value=m.parent_control)
            ws.cell(row=row, column=5, value=m.score)
            ws.cell(row=row, column=6, value=m.confidence.value)
            row += 1

    _apply_confidence_fill(ws, 6, 2, row - 1)
    _auto_width(ws, len(headers), row - 1)


# ── V2 JSON Builder ────────────────────────────────────────────────────────

_NON_AGENTIC_PATTERNS = re.compile(
    r"data policy|regulatory|quality management|processing location|"
    r"cloud vs on-prem|internal process|change approval|"
    r"transparency policy|transparency report",
    re.IGNORECASE,
)


def _gap_reason(ctrl_id: str, title: str) -> str:
    """Determine gap reason for an unmapped control."""
    func = AIUC_CONTROL_FUNCTIONS.get(ctrl_id, "GOVERN")
    if func in ("GOVERN", "DISCLOSE") and _NON_AGENTIC_PATTERNS.search(title):
        return "No agentic-security-specific relevance identified"
    return "Below mapping threshold for all OWASP entries"


def _build_v2_output(mapping: MappingOutput) -> CrosswalkOutput:
    """Transform a v1 MappingOutput into a v2 CrosswalkOutput."""
    w = mapping.metadata.weights
    t = mapping.metadata.thresholds

    # ── source_to_owasp ─────────────────────────────────────────────────
    source_entries: list[SourceControlEntry] = []
    all_rationale_dist: dict[str, int] = {}

    for ctrl in mapping.control_level.aiuc_to_owasp:
        func_class = AIUC_CONTROL_FUNCTIONS.get(ctrl.aiuc_id, "GOVERN")
        v2_mappings: list[SourceToOwaspMapping] = []
        for m in ctrl.mappings:
            code = m.rationale_code.value
            all_rationale_dist[code] = all_rationale_dist.get(code, 0) + 1
            v2_mappings.append(SourceToOwaspMapping(
                owasp_id=m.owasp_id,
                owasp_title=m.owasp_title,
                relevance=m.relevance,
                rationale_code=m.rationale_code,
                rationale_label=m.rationale_label,
                score=m.score,
                signals={
                    "reference_bridge": m.signals.reference_bridge,
                    "semantic": m.signals.semantic,
                    "keyword": m.signals.keyword,
                } if m.signals else None,
            ))

        primary = sum(1 for vm in v2_mappings if vm.relevance == RelevanceLevel.PRIMARY)
        source_entries.append(SourceControlEntry(
            control_id=ctrl.aiuc_id,
            control_title=ctrl.aiuc_title,
            domain=ctrl.aiuc_domain,
            function_class=RationaleCode(func_class),
            mapping_count=len(v2_mappings),
            primary_count=primary,
            secondary_count=len(v2_mappings) - primary,
            mappings=v2_mappings,
        ))

    # ── owasp_to_source ─────────────────────────────────────────────────
    owasp_entries: list[V2OwaspEntry] = []

    for entry in mapping.control_level.owasp_to_aiuc:
        v2_mappings_o: list[OwaspToSourceMapping] = []
        coverage: dict[str, int] = {}

        for m in entry.mappings:
            func = AIUC_CONTROL_FUNCTIONS.get(m.aiuc_id, "GOVERN")
            code = m.rationale_code.value
            coverage[code] = coverage.get(code, 0) + 1

            v2_mappings_o.append(OwaspToSourceMapping(
                control_id=m.aiuc_id,
                control_title=m.aiuc_title,
                domain=m.aiuc_domain,
                function_class=RationaleCode(func),
                relevance=m.relevance,
                rationale_code=m.rationale_code,
                rationale_label=m.rationale_label,
                score=m.score,
                signals={
                    "reference_bridge": m.signals.reference_bridge,
                    "semantic": m.signals.semantic,
                    "keyword": m.signals.keyword,
                } if m.signals else None,
            ))

        uncovered = sorted(
            [RationaleCode(c) for c in RationaleCode if c.value not in coverage],
            key=lambda x: x.value,
        )
        primary = sum(1 for vm in v2_mappings_o if vm.relevance == RelevanceLevel.PRIMARY)

        owasp_entries.append(V2OwaspEntry(
            owasp_id=entry.owasp_id,
            owasp_title=entry.owasp_title,
            owasp_rank=entry.owasp_rank,
            mapping_count=len(v2_mappings_o),
            primary_count=primary,
            secondary_count=len(v2_mappings_o) - primary,
            function_coverage=coverage,
            uncovered_functions=uncovered,
            mappings=v2_mappings_o,
        ))

    # ── summary ─────────────────────────────────────────────────────────
    total_controls = len(source_entries)
    mapped_count = sum(1 for s in source_entries if s.mapping_count > 0)
    total_mappings = sum(s.mapping_count for s in source_entries)
    primary_total = sum(s.primary_count for s in source_entries)
    owasp_with_matches = sum(1 for o in owasp_entries if o.mapping_count > 0)

    summary = CrosswalkSummary(
        controls_mapped=f"{mapped_count} / {total_controls}",
        controls_unmapped=total_controls - mapped_count,
        owasp_entries_with_matches=owasp_with_matches,
        total_mappings=total_mappings,
        primary_mappings=primary_total,
        secondary_mappings=total_mappings - primary_total,
        rationale_distribution=all_rationale_dist,
    )

    # ── gap_analysis ────────────────────────────────────────────────────
    unmapped: list[UnmappedControl] = []
    for s in source_entries:
        if s.mapping_count == 0:
            unmapped.append(UnmappedControl(
                control_id=s.control_id,
                control_title=s.control_title,
                domain=s.domain,
                function_class=s.function_class,
                gap_reason=_gap_reason(s.control_id, s.control_title),
            ))

    gap = GapAnalysis(
        unmapped_controls=unmapped,
        unmapped_count=len(unmapped),
        note=(
            "Unmapped controls did not reach the Related threshold "
            "for any OWASP Agentic Top 10 entry under the current "
            "4-signal scoring model."
        ),
    )

    # ── metadata ────────────────────────────────────────────────────────
    metadata = CrosswalkMetadata(
        schema_version="2.0",
        generated_at=datetime.now(UTC).isoformat(),
        methodology="multi-signal-hybrid-v2",
        source_standard=StandardInfo(
            name="AIUC-1",
            version="1.0",
            url="https://www.aiuc-1.com",
            controls=total_controls,
            domains=len({s.domain for s in source_entries if s.domain}),
        ),
        target_standard=StandardInfo(
            name="OWASP Top 10 for Agentic Applications",
            version="2026",
            url="https://owasp.org/www-project-top-10-for-large-language-model-applications/",
            entries=len(owasp_entries),
        ),
        pipeline=PipelineConfig(
            signals={
                "reference_bridge": PipelineSignal(
                    weight=w.reference_bridge,
                    technique="Jaccard similarity on shared OWASP LLM Top 10 references",
                ),
                "semantic": PipelineSignal(
                    weight=w.semantic,
                    technique=(
                        "Sentence-transformer cosine similarity (all-MiniLM-L6-v2) "
                        "with prevention-guideline boosting and Z-score normalization"
                    ),
                ),
                "keyword": PipelineSignal(
                    weight=w.keyword,
                    technique="TF-IDF cosine similarity with domain-specific synonym expansion",
                ),
                "function_match": PipelineSignal(
                    weight=w.function_boost,
                    technique=(
                        "Multiplicative boost (not additive): "
                        "content * (1 + boost) when control function "
                        "class is in threat function profile"
                    ),
                ),
            },
            thresholds=PipelineThresholds(
                direct=t.direct,
                related=t.related,
                governance_floor=t.governance_floor,
                tangential=t.tangential,
            ),
        ),
        classification=ClassificationConfig(
            rationale_taxonomy=dict(RATIONALE_LABELS),
            relevance_levels=["Primary", "Secondary"],
            classification_accuracy={
                "training": ClassificationAccuracy(
                    standard="AIUC-1",
                    mappings=119,
                    rationale_accuracy=1.0,
                    relevance_accuracy=0.992,
                ),
                "validation": ClassificationAccuracy(
                    standard="NIST AI RMF 1.0",
                    mappings=66,
                    rationale_accuracy=1.0,
                    relevance_accuracy=0.848,
                ),
            },
        ),
    )

    return CrosswalkOutput(
        metadata=metadata,
        summary=summary,
        control_level=V2ControlLevel(
            source_to_owasp=source_entries,
            owasp_to_source=owasp_entries,
        ),
        gap_analysis=gap,
    )


# ── Public API ───────────────────────────────────────────────────────────────


def write_excel(mapping: MappingOutput, output_path: str | Path) -> Path:
    """Write the 5-sheet Excel workbook (README + 4 data sheets)."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _write_readme_sheet(wb)
    _write_aiuc_to_owasp_control(wb, mapping)
    _write_owasp_to_aiuc_control(wb, mapping)
    _write_aiuc_to_owasp_sub(wb, mapping)
    _write_owasp_to_aiuc_sub(wb, mapping)

    wb.save(str(path))
    logger.info("Excel written to %s", path)
    return path


def write_json(mapping: MappingOutput, output_path: str | Path) -> Path:
    """Write the v2 schema-conformant JSON output.

    Validates against crosswalk-mapping-v2.schema.json before writing.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    v2 = _build_v2_output(mapping)
    data = json.loads(v2.model_dump_json(exclude_none=True))

    # Validate against schema
    schema_path = (
        Path(__file__).resolve().parent.parent
        / "schemas"
        / "crosswalk-mapping-v2.schema.json"
    )
    if schema_path.exists():
        import jsonschema

        with open(schema_path) as f:
            schema = json.load(f)
        try:
            jsonschema.validate(data, schema)
            logger.info("JSON output validated against v2 schema")
        except jsonschema.ValidationError as e:
            logger.error("Schema validation failed: %s", e.message)
            print(f"Schema validation FAILED: {e.message}")
            print(f"  Path: {list(e.absolute_path)}")
            raise

    with open(path, "w") as f:
        json.dump(data, f, indent=2)

    logger.info("JSON written to %s", path)
    return path


def generate_outputs(
    mapping: MappingOutput,
    output_dir: str | Path = "mapping",
) -> tuple[Path, Path]:
    """Generate both Excel and JSON outputs.

    Returns:
        Tuple of (excel_path, json_path).
    """
    out = Path(output_dir)
    excel_path = write_excel(mapping, out / "aiuc_owasp_mapping.xlsx")
    json_path = write_json(mapping, out / "aiuc_owasp_mapping.json")
    return excel_path, json_path

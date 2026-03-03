"""Tests for the README sheet feature in aiuc/output.py.

Generates a workbook from a minimal MappingOutput fixture, then opens
it with openpyxl to verify the README sheet structure, content, styling,
and that existing data sheets remain intact.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook import Workbook

from aiuc.models import MappingOutput
from aiuc.output import write_excel

EXPECTED_SHEETS: list[str] = [
    "README",
    "AIUC\u2192OWASP (Control)",
    "OWASP\u2192AIUC (Control)",
    "AIUC\u2192OWASP (Activity)",
    "OWASP\u2192AIUC (Activity)",
]

EXPECTED_SECTION_HEADERS: list[str] = [
    "Workbook Overview",
    "Tab-by-Tab Guide",
    "Understanding Confidence Tiers",
    "Relationship Types",
    "Scoring Methodology",
    "AIUC-1 Domains Quick Reference",
    "OWASP Agentic Top 10 (2026) Quick Reference",
    "How to Get Started",
]

# Expected first header cell (row 1, column 1) for each data sheet.
EXPECTED_DATA_SHEET_HEADERS: dict[str, str] = {
    "AIUC\u2192OWASP (Control)": "AIUC ID",
    "OWASP\u2192AIUC (Control)": "OWASP ID",
    "AIUC\u2192OWASP (Activity)": "Activity ID",
    "OWASP\u2192AIUC (Activity)": "OWASP ID",
}


@pytest.fixture(scope="module")
def workbook() -> Workbook:
    """Generate an Excel workbook from an empty MappingOutput and return it."""
    mapping = MappingOutput()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "test_output.xlsx"
        write_excel(mapping, path)
        wb = openpyxl.load_workbook(str(path))
    return wb


class TestReadmeSheetStructure:
    """Verify that the README sheet exists and is positioned correctly."""

    def test_readme_is_first_sheet(self, workbook: Workbook) -> None:
        """README must be the first tab in the workbook."""
        assert workbook.sheetnames[0] == "README"

    def test_workbook_has_five_sheets(self, workbook: Workbook) -> None:
        """Workbook must contain exactly 5 sheets in the expected order."""
        assert workbook.sheetnames == EXPECTED_SHEETS


class TestReadmeContent:
    """Verify that the README sheet contains all required section headers."""

    def test_readme_contains_sections(self, workbook: Workbook) -> None:
        """All expected section header strings must appear in cell values."""
        ws = workbook["README"]
        all_values: set[str] = set()
        for row in ws.iter_rows(values_only=True):
            for cell_value in row:
                if cell_value is not None:
                    all_values.add(str(cell_value))

        for header in EXPECTED_SECTION_HEADERS:
            assert header in all_values, (
                f"Section header '{header}' not found in README sheet"
            )


class TestReadmeStyling:
    """Verify color-coding and column widths on the README sheet."""

    def test_confidence_tier_colors(self, workbook: Workbook) -> None:
        """Cells labelled 'Green' / 'Yellow' must have the correct fill in column B."""
        ws = workbook["README"]
        green_found = False
        yellow_found = False

        for row in ws.iter_rows(min_col=1, max_col=4):
            cell_b = row[1]

            if cell_b.value == "Green":
                green_found = True
                fill_color = cell_b.fill.start_color.rgb
                # openpyxl may prefix with "00" alpha channel
                assert fill_color in ("C6EFCE", "00C6EFCE"), (
                    f"Expected Direct/green fill C6EFCE, got {fill_color}"
                )

            if cell_b.value == "Yellow":
                yellow_found = True
                fill_color = cell_b.fill.start_color.rgb
                assert fill_color in ("FFEB9C", "00FFEB9C"), (
                    f"Expected Related/yellow fill FFEB9C, got {fill_color}"
                )

        assert green_found, "No cell with value 'Green' found in README column B"
        assert yellow_found, "No cell with value 'Yellow' found in README column B"

    def test_readme_column_widths(self, workbook: Workbook) -> None:
        """Columns A through D on the README sheet must each have width > 10."""
        ws = workbook["README"]
        for col_letter in ("A", "B", "C", "D"):
            width = ws.column_dimensions[col_letter].width
            assert width is not None and width > 10, (
                f"Column {col_letter} width ({width}) must be > 10"
            )


class TestDataSheetsUnchanged:
    """Verify the 4 data sheets still have the correct header rows."""

    def test_existing_sheets_unchanged(self, workbook: Workbook) -> None:
        """First cell (A1) of each data sheet must match the expected header."""
        for sheet_name, expected_header in EXPECTED_DATA_SHEET_HEADERS.items():
            ws = workbook[sheet_name]
            actual = ws.cell(row=1, column=1).value
            assert actual == expected_header, (
                f"Sheet '{sheet_name}' A1: expected '{expected_header}', "
                f"got '{actual}'"
            )
